"""
Excel data source — openpyxl + DuckDB.

Reads .xlsx files into DuckDB for SQL querying. Each sheet becomes a table.
Supports multiple files (each file's sheets are merged into the schema).
"""

import hashlib
import re
from pathlib import Path

import duckdb
import openpyxl

from datasources.base import DataSource


def _safe_table_name(file_stem: str, sheet_name: str) -> str:
    """Create a valid SQL table name from file + sheet names."""
    raw = f"{file_stem}__{sheet_name}"
    return re.sub(r"[^a-zA-Z0-9_]", "_", raw)


def _infer_duckdb_type(values: list) -> str:
    """Infer a DuckDB column type from sample values."""
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "VARCHAR"
    sample = non_null[:50]
    if all(isinstance(v, bool) for v in sample):
        return "BOOLEAN"
    if all(isinstance(v, int) for v in sample):
        return "BIGINT"
    if all(isinstance(v, (int, float)) for v in sample):
        return "DOUBLE"
    from datetime import datetime as dt
    if all(isinstance(v, dt) for v in sample):
        return "TIMESTAMP"
    return "VARCHAR"


class ExcelSource(DataSource):
    """
    DataSource backed by one or more Excel (.xlsx) files.

    Uses openpyxl for reading and DuckDB in-process for SQL querying.
    Each sheet in each file becomes a queryable table.

    Usage:
        source = ExcelSource()
        await source.connect(files=[Path("data.xlsx"), Path("lookup.xlsx")])
        result = await source.query("SELECT * FROM data__Sheet1 LIMIT 10")
        await source.disconnect()
    """

    def __init__(self):
        self._db: duckdb.DuckDBPyConnection | None = None
        self._files: list[Path] = []
        self._tables: dict[str, dict] = {}  # table_name → metadata
        self._source_hash = ""

    async def connect(self, **kwargs) -> None:
        files = kwargs.get("files", [])
        if not files:
            raise ValueError("At least one Excel file path is required")

        self._files = [Path(f) for f in files]
        for f in self._files:
            if not f.exists():
                raise FileNotFoundError(f"Excel file not found: {f}")

        # Create in-memory DuckDB
        self._db = duckdb.connect(":memory:")

        # Compute stable source ID from file paths + sizes
        h = hashlib.md5()
        for f in sorted(self._files):
            h.update(f"{f}:{f.stat().st_size}".encode())
        self._source_hash = h.hexdigest()[:16]

        # Load each file's sheets as tables
        self._tables = {}
        for fpath in self._files:
            await self._load_workbook(fpath)

        print(f"[Excel] Loaded {len(self._tables)} table(s) from "
              f"{len(self._files)} file(s)")

    async def _load_workbook(self, fpath: Path) -> None:
        """Load all sheets from a workbook into DuckDB tables."""
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue  # Skip empty or header-only sheets

            # First row = headers
            headers = []
            for i, h in enumerate(rows[0]):
                col_name = str(h).strip() if h else f"col_{i}"
                # Sanitize column name for SQL
                col_name = re.sub(r"[^a-zA-Z0-9_]", "_", col_name)
                if not col_name or col_name[0].isdigit():
                    col_name = f"c_{col_name}"
                headers.append(col_name)

            data_rows = rows[1:]
            table_name = _safe_table_name(fpath.stem, sheet_name)

            # Infer types from data
            col_types = []
            for col_idx in range(len(headers)):
                col_values = [r[col_idx] if col_idx < len(r) else None
                              for r in data_rows]
                col_types.append(_infer_duckdb_type(col_values))

            # Create table
            col_defs = ", ".join(
                f'"{h}" {t}' for h, t in zip(headers, col_types)
            )
            self._db.execute(f'CREATE TABLE "{table_name}" ({col_defs})')

            # Insert data
            placeholders = ", ".join(["?"] * len(headers))
            for row in data_rows:
                values = []
                for col_idx in range(len(headers)):
                    v = row[col_idx] if col_idx < len(row) else None
                    # Convert to string if type mismatch with VARCHAR
                    if col_types[col_idx] == "VARCHAR" and v is not None:
                        v = str(v)
                    values.append(v)
                try:
                    self._db.execute(
                        f'INSERT INTO "{table_name}" VALUES ({placeholders})',
                        values
                    )
                except Exception:
                    pass  # Skip malformed rows

            self._tables[table_name] = {
                "name":         table_name,
                "source_file":  fpath.name,
                "sheet_name":   sheet_name,
                "columns":      headers,
                "column_types": col_types,
                "row_count":    len(data_rows),
            }

        wb.close()

    async def disconnect(self) -> None:
        if self._db:
            self._db.close()
            self._db = None
        self._tables = {}

    async def query(self, query_text: str) -> dict:
        if not self._db:
            return {"success": False, "message": "Not connected"}
        try:
            result = self._db.execute(query_text)
            columns = [desc[0] for desc in result.description]
            rows = []
            for row in result.fetchall():
                row_dict = {}
                for col_name, value in zip(columns, row):
                    # Convert to JSON-safe types
                    if hasattr(value, 'isoformat'):
                        value = value.isoformat()
                    row_dict[f"[{col_name}]"] = value
                rows.append(row_dict)
            return {"success": True, "data": {"rows": rows}}
        except Exception as e:
            return {"success": False, "message": str(e)}

    async def get_schema(self) -> dict:
        tables = []
        for tname, meta in self._tables.items():
            columns = []
            for col_name, col_type in zip(meta["columns"], meta["column_types"]):
                columns.append({
                    "name":        col_name,
                    "data_type":   col_type,
                    "is_nullable": True,
                    "is_hidden":   False,
                })
            tables.append({
                "name":        tname,
                "columns":     columns,
                "row_count":   meta["row_count"],
                "source_file": meta.get("source_file", ""),
                "sheet_name":  meta.get("sheet_name", ""),
                "description": "",
            })

        # No relationships auto-discovered for Excel
        return {"tables": tables, "relationships": []}

    async def get_sample_data(self, table_name: str,
                               max_rows: int = 100) -> list[dict]:
        result = await self.query(
            f'SELECT * FROM "{table_name}" LIMIT {max_rows}'
        )
        rows = result.get("data", {}).get("rows", [])
        # Clean column names: "[col]" → "col"
        clean = []
        for r in rows:
            clean_row = {}
            for k, v in r.items():
                col = k.strip("[]") if k.startswith("[") else k
                clean_row[col] = v
            clean.append(clean_row)
        return clean

    def source_type(self) -> str:
        return "excel"

    def source_id(self) -> str:
        return f"excel:{self._source_hash}" if self._source_hash else "excel:not_loaded"

    def display_name(self) -> str:
        """Human-readable name showing the loaded file names."""
        if self._files:
            return ", ".join(f.name for f in self._files)
        return self.source_id()

    def query_language(self) -> str:
        return "SQL"

    def supports_writeback(self) -> bool:
        return False

    def table_names(self) -> list[str]:
        """Return all loaded table names."""
        return list(self._tables.keys())
